"""Wallet data export utilities — XLSX, CSV, and JSON.

Provides a standalone ``WalletExporter`` that takes a list of
``WalletEntry`` objects (optionally filtered through ``AllowlistFilter``)
and writes them to disk in various formats.

XLSX is the primary format for analyst handoff (matching AWH's original
output). CSV and JSON are available for programmatic consumption.
"""

from __future__ import annotations

import csv
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from ssi.wallet.allowlist import AllowlistFilter
from ssi.wallet.models import WalletEntry, WalletHarvest

logger = logging.getLogger(__name__)

# Column headers for tabular output (XLSX / CSV)
HEADERS: list[str] = [
    "site_url",
    "token_label",
    "token_symbol",
    "network_label",
    "network_short",
    "wallet_address",
    "harvested_at",
    "run_id",
    "source",
    "confidence",
]


def _entry_to_row(entry: WalletEntry) -> list[Any]:
    """Convert a ``WalletEntry`` to a flat row matching ``HEADERS``."""
    return [
        entry.site_url,
        entry.token_label,
        entry.token_symbol,
        entry.network_label,
        entry.network_short,
        entry.wallet_address,
        entry.harvested_at.isoformat() if entry.harvested_at else "",
        entry.run_id,
        entry.source,
        entry.confidence,
    ]


class WalletExporter:
    """Export wallet entries to XLSX, CSV, or JSON files.

    Optionally applies allowlist filtering before export.

    Usage::

        exporter = WalletExporter(allowlist_filter=AllowlistFilter.default())
        stats = exporter.to_xlsx(entries, Path("output/wallets.xlsx"))
        stats = exporter.to_csv(entries, Path("output/wallets.csv"))
        stats = exporter.to_json(entries, Path("output/wallets.json"))
    """

    def __init__(self, allowlist_filter: AllowlistFilter | None = None) -> None:
        self._filter = allowlist_filter

    def _apply_filter(self, entries: list[WalletEntry]) -> tuple[list[WalletEntry], list[WalletEntry]]:
        """Apply allowlist filter if configured. Returns (to_export, discarded)."""
        if self._filter:
            return self._filter.filter(entries)
        return entries, []

    # -- XLSX export -------------------------------------------------------

    def to_xlsx(
        self,
        entries: list[WalletEntry],
        output_path: Path,
        *,
        sheet_name: str = "Wallets",
        apply_filter: bool = True,
    ) -> dict[str, Any]:
        """Write wallet entries to an XLSX file.

        Args:
            entries: Wallet entries to export.
            output_path: Destination file path.
            sheet_name: Worksheet name.
            apply_filter: Whether to apply allowlist filtering.

        Returns:
            Dict with export statistics (exported, discarded, path).

        Raises:
            ImportError: If ``openpyxl`` is not installed.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for XLSX export. Install with: pip install openpyxl"
            )

        to_export, discarded = self._apply_filter(entries) if apply_filter else (entries, [])

        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name  # type: ignore[union-attr]

        # Header row with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)  # type: ignore[union-attr]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, entry in enumerate(to_export, 2):
            for col_idx, value in enumerate(_entry_to_row(entry), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)  # type: ignore[union-attr]

        # Auto-fit column widths (approximate)
        for col_idx, header in enumerate(HEADERS, 1):
            max_len = len(header)
            for row_idx in range(2, len(to_export) + 2):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")  # type: ignore[union-attr]
                max_len = max(max_len, len(cell_value))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 2, 50)  # type: ignore[union-attr]

        wb.save(str(output_path))
        logger.info("XLSX export: %d entries → %s", len(to_export), output_path)

        return {
            "format": "xlsx",
            "path": str(output_path),
            "exported": len(to_export),
            "discarded": len(discarded),
            "total": len(entries),
        }

    # -- CSV export --------------------------------------------------------

    def to_csv(
        self,
        entries: list[WalletEntry],
        output_path: Path,
        *,
        apply_filter: bool = True,
    ) -> dict[str, Any]:
        """Write wallet entries to a CSV file.

        Args:
            entries: Wallet entries to export.
            output_path: Destination file path.
            apply_filter: Whether to apply allowlist filtering.

        Returns:
            Dict with export statistics.
        """
        to_export, discarded = self._apply_filter(entries) if apply_filter else (entries, [])

        output_path.parent.mkdir(parents=True, exist_ok=True)

        with output_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(HEADERS)
            for entry in to_export:
                writer.writerow(_entry_to_row(entry))

        logger.info("CSV export: %d entries → %s", len(to_export), output_path)

        return {
            "format": "csv",
            "path": str(output_path),
            "exported": len(to_export),
            "discarded": len(discarded),
            "total": len(entries),
        }

    # -- JSON export -------------------------------------------------------

    def to_json(
        self,
        entries: list[WalletEntry],
        output_path: Path,
        *,
        apply_filter: bool = True,
        indent: int = 2,
    ) -> dict[str, Any]:
        """Write wallet entries to a JSON file.

        Args:
            entries: Wallet entries to export.
            output_path: Destination file path.
            apply_filter: Whether to apply allowlist filtering.
            indent: JSON indentation level.

        Returns:
            Dict with export statistics.
        """
        to_export, discarded = self._apply_filter(entries) if apply_filter else (entries, [])

        output_path.parent.mkdir(parents=True, exist_ok=True)

        data = {
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "count": len(to_export),
            "discarded_count": len(discarded),
            "entries": [e.to_dict() for e in to_export],
        }
        output_path.write_text(json.dumps(data, indent=indent), encoding="utf-8")

        logger.info("JSON export: %d entries → %s", len(to_export), output_path)

        return {
            "format": "json",
            "path": str(output_path),
            "exported": len(to_export),
            "discarded": len(discarded),
            "total": len(entries),
        }


def export_harvest(
    harvest: WalletHarvest,
    output_dir: Path,
    *,
    formats: list[str] | None = None,
    allowlist_filter: AllowlistFilter | None = None,
) -> list[dict[str, Any]]:
    """Export a ``WalletHarvest`` to one or more formats.

    Convenience function that wraps ``WalletExporter`` and names output files
    based on the harvest's ``run_id``.

    Args:
        harvest: The wallet harvest to export.
        output_dir: Directory for output files.
        formats: List of formats to export to. Defaults to ``["xlsx", "json"]``.
        allowlist_filter: Optional allowlist filter to apply.

    Returns:
        List of export statistics dicts.
    """
    if formats is None:
        formats = ["xlsx", "json"]

    exporter = WalletExporter(allowlist_filter=allowlist_filter)
    prefix = harvest.run_id or "wallets"
    results: list[dict[str, Any]] = []

    for fmt in formats:
        output_path = output_dir / f"{prefix}.{fmt}"
        if fmt == "xlsx":
            results.append(exporter.to_xlsx(harvest.entries, output_path))
        elif fmt == "csv":
            results.append(exporter.to_csv(harvest.entries, output_path))
        elif fmt == "json":
            results.append(exporter.to_json(harvest.entries, output_path))
        else:
            logger.warning("Unknown export format: %s — skipping", fmt)

    return results
